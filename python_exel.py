from openpyxl import load_workbook


wb = load_workbook("kassenbuch.xlsx")
ws = wb.active

ws['C1'] = "Montag"
ws['E1'] = "29.11.2025"

ws['A4'] = "09:34"
ws['B4'] = "50€"
ws['C4'] = "507206"
ws['D4'] = "603465"
ws['E4'] = "Ladybug"
ws['C5'] = "Ladybug"
ws['B6'] = "+10€ Schulden Ladybug"
ws['B7'] = "-5€ Brötchen Ladybug"
ws['B8'] = "+5€ Brötchen Ladybug"
ws['B9'] = "-30€ Einkauf Rewe Snake"
ws['D6'] = "+30€ Schulden Snake"
#ws['D7'] 
#ws['D8']
#ws['D9']

# y+7 für jede Schicht und 6 Mal möglich

# Endabrechnung
#ws['C53']
#ws['C54']
#ws['C55']
#ws['C56']
#ws['C57']
#ws['C58']
#ws['E58']


wb.save('test.xlsx')