from openpyxl import load_workbook
import locale, datetime
locale.setlocale(locale.LC_TIME, locale.normalize("de"))
import sqlite3
con = sqlite3.connect('Schuldenliste.db')
cur = con.cursor()
# just for debugging
from time import sleep

# interne Funktionen

def KeyMitglied(Vorname, Nachname):
    # gebe den key eines Mitglieds anhand Klarnamens zurück
    query = "SELECT ID FROM User WHERE Vorname = '" + str(Vorname) + "' AND Nachname = '" + str(Nachname) + "';"
    normals = cur.execute(query)
    X = normals.fetchone()
    if X == None:
        return False
    else: 
        return X[0]    
    
def KeyAlias(Alias):
    # gebe den key eines Mitglieds anhand Alias zurück
    query = "SELECT ID FROM User WHERE Alias ='" + str(Alias) + "';"
    normals = cur.execute(query)
    X = normals.fetchone()
    if X == None:
        return False
    else: 
        return X[0]  

def getKey(Vorname, Nachname):
    # beide Möglichkeiten zum Key bekommen 
    if Nachname != None:
        MitgliedKey = KeyMitglied(Vorname, Nachname)
    else:
        MitgliedKey = KeyAlias(Vorname)
    if MitgliedKey == False:
        print("Rechtschreibfehler im Namen oder Alias")
        pass
    return MitgliedKey

def Verantwortlichkeitsliste():
    # alle Verantwortlichen von  Öffnung bis Schließung
    query = "SELECT * FROM Uebergabe WHERE ID >= (SELECT MAX(ID) ID FROM Uebergabe WHERE BisherVerantwortlich = 0);"
    normals = cur.execute(query) 
    Keys = normals.fetchall()
    return Keys


# nochmal checken
def Bestandscheck(Geldbetrag, WeißeKE, BlaueKE):
    # letzte Bestandsmeldung
    query = "SELECT Barbestand, KEWeiß, KEBlau FROM Kassenstand WHERE ID = (SELECT MAX(ID) FROM Kassenstand);"
    normals = cur.execute(query)
    letzterStand = normals.fetchone()
    # Summe aller Bons in der Zwischenzeit
    query = "SELECT SUM(Betrag) FROM (BONS, (SELECT Datum D0, Zeitpunkt T0 FROM Kassenstand WHERE ID = (SELECT MAX(ID) FROM Kassenstand))) WHERE Datum >= D0 AND Zeitpunkt >= T0;"
    normals = cur.execute(query)
    BonSumme = normals.fetchone()[0]
    # falls keine Bons in der Zwischenzeit
    if BonSumme == None:
        BonSumme = 0
    # Schuldeneinzahlungen in der Zwischenzeit
    query = "SELECT SUM(Betrag) FROM (Schuldenliste, (SELECT Datum D0, Zeitpunkt T0 FROM Kassenstand WHERE ID = (SELECT MAX(ID) FROM Kassenstand))) WHERE Datum >= D0 AND Zeitpunkt >= T0 AND Betrag < 0;"
    normals = cur.execute(query)
    SchuldenSumme = normals.fetchone()[0]
    # falls keine Schulden in der Zwischenzeit
    if SchuldenSumme == None:
        SchuldenSumme = 0
    # Abrechnung
    Fehlbetrag = (Geldbetrag - letzterStand[0]) + (BonSumme + SchuldenSumme) - 0.25*(WeißeKE - letzterStand[1]) - 0.5*(BlaueKE - letzterStand[2])
    return Fehlbetrag
    

def Bestandsmeldung(Geldbetrag, WeißeKE, BlaueKE, Grund):
    # Abfrage jetzt veantwortlich (key)
    aktuellKey = Verantwortlich() 
    Fehlbetrag = Bestandscheck(Geldbetrag, WeißeKE, BlaueKE)
    # checke ob Abrechnung mit allen Veränderungen passt - Fehlerbehebung???
    if Fehlbetrag != 0:
        # Fehlbetrag - bisher nicht implementiert
        print('Der Kassenstand stimmt nicht mit dem erwarteten Betrag überin. Fehlbetrag: ' + str(Fehlbetrag))
        pass
        # trotzdem einfügen? - über den ganzen Öffnungszyklus checken? 
    # Bestandsmeludng einfügen
    Values = "('" + str(Grund) + "', " + str(aktuellKey) + ", " + str(Geldbetrag) + ", " + str(WeißeKE) + ", " + str(BlaueKE) + ");"
    query = "INSERT INTO Kassenstand (Art, Verantwortlich, Barbestand, KEWeiß, KEBlau) VALUES" + Values
    cur.execute(query)
    con.commit()

def Uebergabe(altverant, neuverant):
    query = "INSERT INTO Uebergabe (BisherVerantwortlich, Verantwortlich) VALUES (" + str(altverant) + ", " + str(neuverant) + ");"
    cur.execute(query)
    con.commit()

def getName(key):
    query = "SELECT Alias FROM User WHERE ID ='" + str(key) + "';"
    normals = cur.execute(query)
    X = normals.fetchone()[0]
    if X != None:
        return X
    query = "SELECT Vorname, Nachname FROM User WHERE ID ='" + str(key) + "';"
    normals = cur.execute(query)
    X = normals.fetchone()
    return X[0]+' '+X[1]


# Probleme: bisher nur eine Exel-Tablelle mur max 6 Schichtübergaben + Schließung pro Tag und pro Schicht max 8 Umsatzeinträge
def Kassenbuch():
    # alle Übergaben während Öffnung
    query = "SELECT Datum, Zeitpunkt, Barbestand, KEWeiß, KEBlau, Verantwortlich FROM Kassenstand WHERE ID>=(SELECT MAX(ID) FROM Kassenstand WHERE Art = 'Öffnung')"
    normals = cur.execute(query)
    Uebergaben = normals.fetchall()
    # alle Verantwotlichen während der Öffnung
    query = "SELECT Verantwortlich FROM Uebergabe WHERE ID >= (SELECT MAX(ID) FROM Uebergabe WHERE BisherVerantwortlich = 0)"
    normals = cur.execute(query)
    Verantwortlich = normals.fetchall()
    # Bearbeitung der Tabelle
    wb = load_workbook("kassenbuch.xlsx")
    ws = wb.active 
    day = Uebergaben[0][0][8:]
    month = Uebergaben[0][0][5:7] 
    year = Uebergaben[0][0][:4]
    x = datetime.datetime(int(year), int(month), int(day))
    ws['C1'] = x.strftime("%A") 
    ws['E1'] = day+'.'+month+'.'+year
    if len(Uebergaben)-1 >= 6:
        print("zu viele Übergaben - aktuell nicht implementiert")  
    for handover in range(len(Uebergaben)-1):
        # Zeitpunkt der Öffnung/Übergabe
        ws['A'+ str(4+7*handover)] = Uebergaben[handover][1][:5]
        # Kassenstand zur Öffnung/Übergabe
        ws['B'+ str(4+7*handover)] = str(Uebergaben[handover][2])+'€'
        # Markenummer weiß zur Öffnung/Übergabe
        ws['C'+ str(4+7*handover)] = Uebergaben[handover][3]
        # Markenummer blau zur Öffnung/Übergabe
        ws['D'+ str(4+7*handover)] = Uebergaben[handover][4]
        # Öffnung/Übergabe gezählt von
        ws['E'+ str(4+7*handover)] = getName(Uebergaben[handover][5])
        # ab jetzt verantwortlich
        ws['C'+ str(5+7*handover)] = getName(Verantwortlich[handover][0])
        # alle Bons in dem Zeitraum
        query = "SELECT Mitglied, Betrag, Zweck FROM Bons WHERE Datum >= '" + Uebergaben[handover][0] + "' AND Zeitpunkt >= '" + Uebergaben[handover][1] + "' AND Datum <= '" + Uebergaben[handover+1][0] + "' AND Zeitpunkt < '" + Uebergaben[handover+1][1] + "'"
        normals = cur.execute(query)
        Bons = normals.fetchall()
        # alle bezahlten Schulden in dem Zeitraum
        query = "SELECT Mitglied, Betrag FROM Schuldenliste WHERE Datum >= '" + Uebergaben[handover][0] + "' AND Zeitpunkt >= '" + Uebergaben[handover][1] + "' AND Datum <= '" + Uebergaben[handover+1][0] + "' AND Zeitpunkt < '" + Uebergaben[handover+1][1] + "' AND Betrag < 0"
        normals = cur.execute(query)
        Schulden = normals.fetchall()        
        # alle Umsätze in dem Zeitraum
        Umsaetze = []
        for x in range(len(Bons)):
            for y in Schulden:
                if Bons[x][0] == y[0] and Bons[x][1] == -y[1]:
                    Umsaetze.append((Bons[x][0], '±'+str(Bons[x][1]), Bons[x][2]+"/Schulden"))
                    break
            if len(Umsaetze) <= x:
                Umsaetze.append(Bons[x])
        for y in Schulden:
            done = 0
            for x in Umsaetze:
                if y[0] == x[0] and type(x[1]) == str and float(x[1][1:]) == -y[1]:
                    done = 1
                    break
            if done == 0:
                Umsaetze.append((y[0], "+" + str(-1*y[1]), "Schulden"))
        # Einfügen der Umsätze in die Tabelle
        if len(Umsaetze) >= 9:
            print("zu viele Umsätze - aktuell nicht implementiert")            
        for trans in range(len(Umsaetze)):
            Mitglied = getName(Umsaetze[trans][0])
            if Umsaetze[trans][2] == 'Schlitz':
                Eintrag = "-"+str(Umsaetze[trans][1]) +"€ " + Umsaetze[trans][2]
            else:
                Eintrag = str(Umsaetze[trans][1]) +"€ " + Umsaetze[trans][2] + " " + Mitglied
            if trans < 4:
                ws['B'+ str(6+trans+7*handover)] = Eintrag
            else:
                ws['D'+ str(6+trans-4+7*handover)] = Eintrag
    # Tagesabschluss
    # Zeitpunkt der Schließung
    ws['A53'] = Uebergaben[-1][1][:5]
    # Kassenstand zur Schließung
    ws['C58'] = str(Uebergaben[-1][2])+'€'
    # Markenummer weiß zur Schließung
    ws['C54'] = Uebergaben[-1][3]
    # Markenummer blau zur Schließung
    ws['C55'] = Uebergaben[-1][4]
    # Abrechnung gemacht von
    ws['E58'] = getName(Uebergaben[-1][5])
    # Markenumsätze
    ws['E54'] = (ws['C54'].value - ws['C4'].value)/4 + (ws['C55'].value - ws['D4'].value)/2
    # Summe der Bons während Öffnung
    query = "SELECT SUM(Betrag) FROM Bons WHERE Datum >= (SELECT Datum FROM Kassenstand WHERE ID = (SELECT MAX(ID) FROM Kassenstand WHERE Art = 'Öffnung')) AND Zeitpunkt >= (SELECT Zeitpunkt FROM Kassenstand WHERE ID = (SELECT MAX(ID) FROM Kassenstand WHERE Art = 'Öffnung'))"
    normals = cur.execute(query)
    Bonsumme = normals.fetchone()
    # Summe der bezahlten Schulden während Öffnung
    query = "SELECT SUM(Betrag) FROM Schuldenliste WHERE Datum >= (SELECT Datum FROM Kassenstand WHERE ID = (SELECT MAX(ID) FROM Kassenstand WHERE Art = 'Öffnung')) AND Zeitpunkt >= (SELECT Zeitpunkt FROM Kassenstand WHERE ID = (SELECT MAX(ID) FROM Kassenstand WHERE Art = 'Öffnung')) AND Betrag < 0"
    normals = cur.execute(query)
    Schuldensumme = normals.fetchone()
    ws['E55'] = -1*(Bonsumme[0]+Schuldensumme[0])
    # Gesmtsumme der Umsätze
    ws['E56'] = ws['E54'].value + ws['E55'].value
    wb.save('test.xlsx')




# Member-Amt Funktionen

def neuesMitglied(Vorname, Nachname):
    # füge ein neues Mitglied hinzu
    query = "SELECT ID FROM User WHERE Vorname = '" + str(Vorname) + "' AND Nachname = '" + str(Nachname) + "';"
    normals = cur.execute(query)
    if normals.fetchone() == None:
        # erstellung eines neuen Mitglieds
        query = "INSERT INTO User (Vorname, Nachname) VALUES ('" + str(Vorname) + "', '" + str(Nachname) + "');"
        cur.execute(query)
        con.commit()
    else:
        # Mitglied-Name bereits vergeben - aktuell irrelevant und nicht implementiert
        print("Das Mitglied ist bereis im System - bitte einen eindeutigen Namen verwenden.")
        pass
    
def AliasAendern(Vorname, Nachname, Alias):
    # füge ein Alias für das Mitglied hinzu/ändert ihn -> was wir heute auf der Liste benutzen
    MitgliedKey = KeyMitglied(Vorname, Nachname)    
    # checke Alias bereits belegt
    query = "SELECT ID FROM User WHERE Alias ='" + str(Alias) + "';"
    normals = cur.execute(query)
    X = normals.fetchone()
    if X != None and X[0] != MitgliedKey:
        # alias ist bereits von einem anderen Mitgleid belegt
        print("der Alias ist bereits vergeben. Bitte wähle eoinen anderen alias")
        pass
    else:
        # setzt den Alias des Mitglieds (egal, ob der bereits so war)
        query = "UPDATE User SET Alias = '" + str(Alias) + "' WHERE ID = " + str(MitgliedKey) + ";" 
        cur.execute(query)
        con.commit()

def NamenAendern(altVorname, altNachname, neuVorname, neuNachname):
    # ändert den Klarnamen eines Mitglieds
    MitgliedKey = KeyMitglied(altVorname, altNachname)
    query = "UPDATE User SET Vorname = '" + str(neuVorname) +"', Nachname = '" + str(neuNachname) + "' WHERE ID = " + str(MitgliedKey) +";"
    cur.execute(query)
    con.commit()


def Verantwortlich(Date = None, Zeitpunkt = None):
    # verantwortliche Person zum Zeitpunkt datetime - default aktuell
    # date,time format in sqlite: YYYY-MM-DD, HH:MM:SS - jeweils als string
    if Date == None and Zeitpunkt == None:
        # standardabfrage zur aktuellen Verantwortung
        query = "SELECT Verantwortlich FROM Uebergabe WHERE ID = (SELECT MAX(ID) FROM Uebergabe);"
        normals = cur.execute(query) 
        VerantwortlichKey = normals.fetchone()[0]
    else:
        # Verantwortlich zu einem bestimmten Zeitpunkt
        allbefore = "(SELECT * FROM Uebergabe WHERE Datum <= '" + str(Date) + "' AND Zeitpunkt <= '" + str(Zeitpunkt) + "')"
        query = "SELECT Verantwortlich FROM " + allbefore + " WHERE ID = (SELECT MAX(ID) FROM " + allbefore + ");"
        normals = cur.execute(query) 
        VerantwortlichKey = normals.fetchone()[0]
    return VerantwortlichKey


def orangeSetzen():
    # alle Mitglieder mit >= 5 Schulden werden orange gesetzt
    query = "UPDATE User SET Rot = 1 WHERE ID IN (SELECT Mitglied FROM (SELECT Mitglied, Sum(Betrag) Schulden FROM Schuldenliste GROUP BY Mitglied) WHERE Schulden >= 5);"
    cur.execute(query)
    con.commit()


def rotSetzen():
    # setzt alle orangenen Mitglieder auf rot (kann keine Striche mehr machen)
    query = "UPDATE User SET Rot = 2 WHERE ID IN (SELECT Mitglied FROM (SELECT Mitglied, Sum(Betrag) Schulden FROM Schuldenliste GROUP BY Mitglied) WHERE Schulden >= 5 INTERSECT SELECT ID FROM User WHERE Rot = 1);"
    cur.execute(query)
    con.commit()



# Mitglieds Funktionen

def Striche(Anzahl, Vorname, Nachname = None):
    # eurobetrag der Anzahl Striche als Schulden bei Name hinzufügen
    # Abfrage jetzt veantwortlich (key)
    aktuellKey = Verantwortlich() 
    # Abfrage Mitglied rot gefärbt - sonst Striche als Schulden
    MitgliedKey = getKey(Vorname, Nachname)
    query = "SELECT Rot FROM User WHERE ID = " + str(MitgliedKey) + ";"
    normals = cur.execute(query) 
    MitgliedRot = normals.fetchone()[0]
    if MitgliedRot == 2:
        # Mitglied darf keine Striche machen -- aktuell nicht implementiert
        Schulden = meineSchulden(Vorname, Nachname)
        print("Das Mitglied ist aktell rot hinterlegt. Aktueller Schuldenbetrag des Mitglieds: " + str(Schulden) + ".")
        pass
    else:
        Betrag = Anzahl * 0.25
        query = "INSERT INTO Schuldenliste (Mitglied, Betrag, Verantwortlich) VALUES (" + str(MitgliedKey) + ", " + str(Betrag) + ", " + str(aktuellKey) + ");"
        cur.execute(query)
        con.commit()


def Schuldenbezahlt(Betrag, Vorname, Nachname = None):
    # Abfrage jetzt veantwortlich (key)
    aktuellKey = Verantwortlich() 
    # -eurobetrag  als Schulden bei Name hinzufügen
    MitgliedKey = getKey(Vorname, Nachname)
    query = "INSERT INTO Schuldenliste (Mitglied, Betrag, Verantwortlich) VALUES (" + str(MitgliedKey) + ", " + str(-Betrag) + ", " + str(aktuellKey) + ");"
    cur.execute(query)
    con.commit()
    # ggf rot entfernen
    query = "SELECT Rot FROM User WHERE ID = " + str(MitgliedKey) + ";"
    normals = cur.execute(query) 
    MitgliedRot = normals.fetchone()[0]
    if MitgliedRot == 1 or MitgliedRot == 2:
        # Abfarge Schulden von MitgliedKey < 5 -> setzt Farbe auf weiß
        query = "SELECT Schulden FROM (SELECT Mitglied, Sum(Betrag) Schulden FROM Schuldenliste GROUP BY Mitglied) WHERE Mitglied = " + str(MitgliedKey) + ";"
        normals = cur.execute(query)
        neuerBetrag = normals.fetchone()[0]
        if neuerBetrag < 5:
            query = "UPDATE User SET Rot = 0 WHERE ID = " + str(MitgliedKey) + ";" 
            cur.execute(query)
            con.commit()


# erst Namen dann Zweck?
def eingereichterBon(Betrag, Zweck = "keine Angabe", Vorname = None, Nachname = None):
    # Abfrage jetzt veantwortlich (key)
    aktuellKey = Verantwortlich() 
    # Bon wird von der Person eingereicht
    if Vorname != None:
        MitgliedKey = getKey(Vorname, Nachname)
    else:
        MitgliedKey = aktuellKey
    Values = "(" + str(MitgliedKey) + ", " + str(Betrag) + ", " +str(aktuellKey) + ", '" +str(Zweck) +  "');"
    query = "INSERT INTO Bons (Mitglied, Betrag, Verantwortlich, Zweck) VALUES " + Values
    cur.execute(query)
    con.commit()


def Schlitzgeld(Betrag):
    # als Bon einführen mit Zweck Schlitz
    eingereichterBon(Betrag, "Schlitz")


# Vorname, Nachname von ab jetzt verantwortlich
def Cafeöffnung(Geldbetrag, WeißeKE, BlaueKE, Vorname, Nachname = None):
    # Öffnung möglich
    query = "SELECT Art FROM Kassenstand WHERE ID = (SELECT MAX(ID) FROM Kassenstand);"
    normals = cur.execute(query)
    lastcheck = normals.fetchone()[0]
    if lastcheck != "Schließung":
        print('Cafe ist bereits offen')
        pass
    # Abfrage key abjetztVaerantwortlich
    neuKey = getKey(Vorname, Nachname)
    # Bestandsmeldung (Öffnung)
    Bestandsmeldung(Geldbetrag, WeißeKE, BlaueKE, "Öffnung") 
    # ändere Verantwortung bei Öffnung
    query = "UPDATE Kassenstand SET Verantwortlich = " + str(neuKey) + " WHERE ID = (SELECT MAX(ID) FROM Kassenstand);"
    cur.execute(query)
    con.commit()
    # Übergabe cafe -> Person
    Uebergabe(0, neuKey)


# Vorname, Nachname von ab jetzt verantwortlich
def Verantwortungsübergabe(Geldbetrag, WeißeKE, BlaueKE, Vorname, Nachname = None):
    # Übergabe möglich
    query = "SELECT Art FROM Kassenstand WHERE ID = (SELECT MAX(ID) FROM Kassenstand);"
    normals = cur.execute(query)
    lastcheck = normals.fetchone()[0]
    if lastcheck == "Schließung":
        print('Cafe muss erst geöffnet werden')
        pass
    # Abfrage jetzt veantwortlich (key)
    altKey = Verantwortlich()
    # Abfrage key abjetztVaerantwortlich
    neuKey = getKey(Vorname, Nachname)
    # Bestandsmeldung (Übergabe)
    Bestandsmeldung(Geldbetrag, WeißeKE, BlaueKE, "Zwischen")
    # Übergabe
    Uebergabe(altKey, neuKey)


def Cafeschließung(Geldbetrag, WeißeKE, BlaueKE):
    # Schließung möglich
    query = "SELECT Art FROM Kassenstand WHERE ID = (SELECT MAX(ID) FROM Kassenstand);"
    normals = cur.execute(query)
    lastcheck = normals.fetchone()[0]
    if lastcheck == "Schließung":
        print('Cafe ist bereits geschlossen')
        pass
    # Abfrage jetzt veantwortlich (key)
    altKey = Verantwortlich()
    # Bestandsmeldung (Schließung)
    Bestandsmeldung(Geldbetrag, WeißeKE, BlaueKE, "Schließung")
    # Übergabe Person -> Cafe
    Uebergabe(altKey, 0)
    # Abrechnung aller Schichten incl Fehler (in pdf?)
    # todo: Fehlermeldungen in der pdf
    Kassenbuch()



def meineSchulden(Vorname, Nachname = None):
    # Ausgabe der Schulden der Person
    key = getKey(Vorname, Nachname)
    query = "SELECT Schulden from (SELECT Mitglied, Sum(Betrag) Schulden FROM Schuldenliste GROUP BY Mitglied) WHERE Mitglied = " + str(key) + ";"
    normals = cur.execute(query) 
    return normals.fetchone()[0]

   
"""
Storno (pro aktion letzte Aktion?)
"""

"""
Was passiert bei Anbruch einer neuen Rolle Marken?
"""

#############
#Demo
"""
q1 = "INSERT INTO Kassenstand (Art, Verantwortlich, Barbestand, KEWeiß, KEBlau) VALUES ('Schließung', 0, 40, 615027, 528993);"
cur.execute(q1)
con.commit()
q2 = "INSERT INTO Uebergabe (BisherVerantwortlich, Verantwortlich) VALUES (0, 0);"
cur.execute(q2)
con.commit()


print(neuesMitglied('Mathe', 'Cafe'))
print(AliasAendern('Mathe', 'Cafe', 'Cafe'))

print(neuesMitglied('Donald', 'Duck'))
print(AliasAendern('Donald', 'Duck', 'Dr. Don'))




print(Striche(4, 'Donald', 'Duck'))
print(Striche(5, 'Dr. Don'))

print(meineSchulden('Dr. Don'))

print(Schuldenbezahlt(5, 'Dr. Don'))
print(meineSchulden('Dr. Don'))



print(neuesMitglied('Mickie', 'Msus'))
print(NamenAendern('Mickie', 'Msus', 'Micky', 'Maus'))

print(neuesMitglied('Daisy', 'Duck'))
print(neuesMitglied('Gustav', 'Gans'))
print(neuesMitglied('Donna', 'Duck'))
print(neuesMitglied('Indiana', 'Goof'))


time.sleep(1)
print(Cafeöffnung(45, 615027, 528993, 'Dr. Don'))
time.sleep(1)
print(eingereichterBon(2, "Brötchen"))
print(Schuldenbezahlt(2, 'Dr. Don'))
print(Schuldenbezahlt(10, 'Gustav', 'Gans'))

time.sleep(1)
print(Verantwortungsübergabe(60, 615035, 528999, 'Micky', 'Maus'))
time.sleep(1)
print(eingereichterBon(65, "Getränke SAP"))
print(Schuldenbezahlt(65, 'Micky', 'Maus'))
print(eingereichterBon(32, "Einkauf REWE", 'Donna', 'Duck'))
print(Schuldenbezahlt(32, 'Donna', 'Duck'))
print(Schuldenbezahlt(10, 'Indiana', 'Goof'))

time.sleep(1)
print(Verantwortungsübergabe(100, 615083, 529035, 'Daisy', 'Duck'))
time.sleep(1)
print(Schuldenbezahlt(50, 'Daisy', 'Duck'))
print(Schlitzgeld(100))

time.sleep(1)
print(Cafeschließung(65, 615107, 529053))
"""

#Kassenbuch()





#print(Striche(4, 'Micky', 'Maus'))
#print(Striche(5, 'Maxi'))
#print(meineSchulden('Maxi'))